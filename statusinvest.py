#encoding: utf-8
from urllib.request import Request, urlopen, urlretrieve
from urllib.error import URLError, HTTPError
from bs4 import BeautifulSoup
import pandas as pd
import json
import requests
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill

from sheet2dict import Worksheet
import os, sys

nome_planilha = "Resultado.xlsx"


def adicionar_ou_criar_planilha(lista_dict_acoes, nome_planilha): #recebe lista de dicionarios
        
    dfFinal = pd.DataFrame(lista_dict_acoes) #transforma dicionarios em dataframe
    dfFinal = dfFinal.reindex(['TICKER', 'Nome', 'Valor Atual do Ativo', 'Min 52 Semanas', 'Min Mês', 'Max 52 Semanas', 'Max Mês', 'Dividend Yeld', 'Dividend Yeld 12 Meses', 'Valorização (12M)', 'Valorização Mês Atual', 'Tipo', 'TAG ALONG', 'Liquidez Média Diária', 'P/L', 'PEG RATIO', 'P/VP', 'EV/EBITDA', 'EV/EBIT', 'P/EBITDA', 'P/EBIT', 'VPA', 'P/ATIVO', 'LPA', 'P/SR', 'P/CAP. GIRO', 'P/ATIVO CIRC. LIQ', 'DÍV. LÍQUIDA/PL', 'DÍV. LÍQUIDA/EBITDA', 'DÍV. LÍQUIDA/EBIT', 'PL/ATIVOS', 'PASSIVOS/ATIVOS', 'LIQ. CORRENTE', 'M. BRUTA', 'M. EBITDA', 'M. EBIT', 'M. LÍQUIDA', 'ROE', 'ROA', 'ROIC', 'GIRO ATIVOS', 'CAGR RECEITAS 5 ANOS', 'CAGR LUCROS 5 ANOS', 'Dividendo 2013', 'Dividendo 2014', 'Dividendo 2015', 'Dividendo 2016', 'Dividendo 2017', 'Dividendo 2018', 'Dividendo 2019', 'Dividendo 2020', 'Dividendo 2021', 'Dividendo 2022', 'Payout Médio', 'Payout Atual', 'Payout Menor Valor', 'Payout Maior Valor', 'Payout Médio 2018', 'Payout Médio 2019', 'Payout Médio 2020', 'Payout Médio 2021', 'Payout Médio 2022', 'Lucro Líquido 2018', 'Lucro Líquido 2019', 'Lucro Líquido 2020', 'Lucro Líquido 2021', 'Lucro Líquido 2022', 'Proventos 2018', 'Proventos 2019', 'Proventos 2020', 'Proventos 2021', 'Proventos 2022', 'Patriônio Líquido', 'Ativos', 'Ativo Circulante', 'Dívida Bruta', 'Disponibilidade', 'Dívida Líquida', 'Valor de Mercado', 'Valor de Firma', 'Nº Total de Papéis', 'Segmento de Listagem', 'Free Float'], axis=1)
    dfFinal.to_excel(nome_planilha, index=False)


def pegar_tickers_planilha(nome_planilha):
        df_planilha = pd.read_excel(nome_planilha, sheet_name='Sheet1') # can also index sheet by name or fetch all sheets
        
        df_planilha['TICKER'] = df_planilha['TICKER'].str.upper()
        # print(df_planilha)
        df_planilha = df_planilha.drop_duplicates(subset='TICKER', keep="first")
        df_planilha = df_planilha.dropna(subset='TICKER')
        lista_tickers = df_planilha['TICKER'].tolist()

        if not lista_tickers:
            raise Exception("Planilha vazia")

        return lista_tickers


def organizar_planilha(nome_planilha):
    writer = pd.ExcelWriter(nome_planilha, engine='openpyxl', if_sheet_exists='replace', mode='a')

    workbook  = writer.book



    sheet = workbook['Sheet1']


    for column_cells in sheet.columns:
        new_column_length = max(len(str(cell.value)) for cell in column_cells)
        new_column_letter = (get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            sheet.column_dimensions[new_column_letter].width = new_column_length*1.23

    for row in range(1,sheet.max_row+1):
        for col in range(1,sheet.max_column+1):
            cell=sheet.cell(row, col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row%2==0:
                start_color='BFBFBF' 
            else :
                start_color='00FFFFFF'
            cell.fill = PatternFill(start_color=start_color, fill_type="solid")

        
    

    writer.close()

    return "Planilha Organizada"

def buscarTIK(tik):
        acao =tik
        url = 'https://statusinvest.com.br/acoes/'+ acao
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/76.0.3809.100 Safari/537.36'}

        api_payout = "https://statusinvest.com.br/acao/payoutresult?code="+ acao
        req_payout = Request(api_payout, headers = headers)

        api_dividendo = f"https://statusinvest.com.br/acao/companytickerprovents?ticker={acao}&chartProventsType=2"
        req_dividendo = Request(api_dividendo, headers = headers)

        try:
            req = Request(url, headers = headers)
            response = urlopen(req)

            res_payout = urlopen(req_payout)
            payout_json = json.loads(res_payout.read())

            res_dividendo = urlopen(req_dividendo)
            dividendo_json = json.loads(res_dividendo.read())

        except HTTPError as e:
            print(e.status, e.reason)

        except URLError as e:
            print(e.reason)

        html = response.read().decode('utf-8')
        soup = BeautifulSoup(html, 'html.parser')

        # print(list(filter(lambda ano: ano['rank'] == 2013, dividendo_json['assetEarningsYearlyModels'])))

        dado = []
        dados = {}

        dados['TICKER'] = soup.find("h1", class_="lh-4").get_text().split()[0]
        dados['Nome'] = soup.find("h1", class_="lh-4").get_text()
        dados['Valor Atual do Ativo'] = soup.find_all("strong", class_="value")[0].get_text()
        dados['Min 52 Semanas'] = soup.find_all("strong", class_="value")[1].get_text()
        dados['Min Mês'] = soup.find_all("span", class_="sub-value")[1].get_text()
        dados['Max 52 Semanas'] = soup.find_all("strong", class_="value")[2].get_text()
        dados['Max Mês'] = soup.find_all("span", class_="sub-value")[2].get_text()
        dados['Dividend Yeld'] = soup.find("strong", class_="value d-block lh-4 fs-4 fw-700").get_text()
        dados['Dividend Yeld 12 Meses'] = soup.find_all("span", class_="sub-value")[3].get_text()
        dados['Valorização (12M)'] = soup.find_all("strong", class_="value")[4].get_text()
        dados['Valorização Mês Atual'] = soup.find_all("b", class_="v-align-middle")[1].get_text()
        dados['Tipo'] = soup.find_all("h3", class_="title m-0 mb-1")[0].next_sibling.next_element.get_text()
        dados['TAG ALONG'] = soup.find_all("span", class_="sub-value legend-tooltip pr-2 d-inline-block")[0].next_sibling.next_element.get_text().replace("\n", "")
        dados['Liquidez Média Diária'] = soup.find_all("span", class_="sub-value legend-tooltip pr-2 d-inline-block")[1].next_sibling.next_element.get_text().replace("\n", "")
        dados['P/L'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[1].find('strong').get_text()
        dados['PEG RATIO'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[2].find('strong').get_text()
        dados['P/VP'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[3].find('strong').get_text()
        dados['EV/EBITDA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[4].find('strong').get_text()
        dados['EV/EBIT'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[5].find('strong').get_text()
        dados['P/EBITDA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[6].find('strong').get_text()
        dados['P/EBIT'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[7].find('strong').get_text()
        dados['VPA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[8].find('strong').get_text()
        dados['P/ATIVO'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[9].find('strong').get_text()
        dados['LPA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[10].find('strong').get_text()
        dados['P/SR'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[11].find('strong').get_text()
        dados['P/CAP. GIRO'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[12].find('strong').get_text()
        dados['P/ATIVO CIRC. LIQ'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[13].find('strong').get_text()
        dados['DÍV. LÍQUIDA/PL'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[14].find('strong').get_text()
        dados['DÍV. LÍQUIDA/EBITDA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[15].find('strong').get_text()
        dados['DÍV. LÍQUIDA/EBIT'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[16].find('strong').get_text()
        dados['PL/ATIVOS'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[17].find('strong').get_text()
        dados['PASSIVOS/ATIVOS'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[18].find('strong').get_text()
        dados['LIQ. CORRENTE'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[19].find('strong').get_text()
        dados['M. BRUTA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[20].find('strong').get_text()
        dados['M. EBITDA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[21].find('strong').get_text()
        dados['M. EBIT'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[22].find('strong').get_text()
        dados['M. LÍQUIDA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[23].find('strong').get_text()
        dados['ROE'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[24].find('strong').get_text()
        dados['ROA'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[25].find('strong').get_text()
        dados['ROIC'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[26].find('strong').get_text()
        dados['GIRO ATIVOS'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[27].find('strong').get_text()
        dados['CAGR RECEITAS 5 ANOS'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[28].find('strong').get_text()
        dados['CAGR LUCROS 5 ANOS'] = soup.find_all("div", class_="d-flex align-items-center justify-between pr-1 pr-xs-2")[29].find('strong').get_text()

        for item in dividendo_json['assetEarningsYearlyModels']:
            if item["rank"] >= 2012:
                dados[f'Dividendo {item["rank"]}'] = item["value"]

        dados['Payout Médio'] = payout_json['avg_F']
        dados['Payout Atual'] = payout_json['actual_F']
        dados['Payout Menor Valor'] = payout_json['minValue_F']
        dados['Payout Maior Valor'] = payout_json['maxValue_F']

        for i, item in enumerate(payout_json['chart']['category']):
            dados[f'Payout Médio {item}'] = payout_json['chart']['series']['percentual'][i]['value_F']

        for i, item in enumerate(payout_json['chart']['category']):
            dados[f'Lucro Líquido {item}'] = payout_json['chart']['series']['lucroLiquido'][i]['value_F']

        for i, item in enumerate(payout_json['chart']['category']):
            dados[f'Proventos {item}'] = payout_json['chart']['series']['proventos'][i]['value_F']

        dados['Patriônio Líquido'] = soup.find_all("h3", class_="title m-0")[14].parent.parent.find("strong").get_text()
        dados['Ativos'] = soup.find_all("h3", class_="title m-0")[15].parent.parent.find("strong").get_text()
        dados['Ativo Circulante'] = soup.find_all("h3", class_="title m-0")[16].parent.parent.find("strong").get_text()
        dados['Dívida Bruta'] = soup.find_all("h3", class_="title m-0")[17].parent.parent.find("strong").get_text()
        dados['Disponibilidade'] = soup.find_all("h3", class_="title m-0")[18].parent.parent.find("strong").get_text()
        dados['Dívida Líquida'] = soup.find_all("h3", class_="title m-0")[19].parent.parent.find("strong").get_text()
        dados['Valor de Mercado'] = soup.find_all("h3", class_="title m-0")[20].parent.parent.find("strong").get_text()
        dados['Valor de Firma'] = soup.find_all("h3", class_="title m-0")[21].parent.parent.find("strong").get_text()
        dados['Nº Total de Papéis'] = soup.find_all("h3", class_="title m-0 legend-tooltip")[0].parent.find("strong").get_text()
        dados['Segmento de Listagem'] = soup.find_all("h3", class_="title m-0")[22].parent.parent.find("strong").get_text()
        dados['Free Float'] = soup.find_all("div", class_="title m-0 legend-tooltip d-flex align-items-center")[0].parent.find("strong").get_text()




        dado.append(dados)
        
        ws = Worksheet()



        # df_planilha = pd.read_excel(nome_planilha, sheet_name='Sheet1') # can also index sheet by name or fetch all sheets
        # df_planilha.drop_duplicates(subset='TICKER', keep="first")

        # dictPlanilha = ws.xlsx_to_dict(path=nome_planilha)


        # dfFinal = pd.concat([df, df_planilha], axis=0)
        # dfPlanilha = pd.DataFrame(data=dictPlanilha.sheet_items)
        # dfFinal = dfFinal.drop_duplicates(subset='TICKER')
        # dfOrdered = dfFinal[['TICKER', 'Nome', 'Valor Atual do Ativo', 'Min 52 Semanas', 'Min Mês', 'Max 52 Semanas', 'Max Mês', 'Dividend Yeld', 'Dividend Yeld 12 Meses', 'Valorização (12M)', 'Valorização Mês Atual', 'Tipo', 'TAG ALONG', 'Liquidez Média Diária', 'P/L', 'PEG RATIO', 'P/VP', 'EV/EBITDA', 'EV/EBIT', 'P/EBITDA', 'P/EBIT', 'VPA', 'P/ATIVO', 'LPA', 'P/SR', 'P/CAP. GIRO', 'P/ATIVO CIRC. LIQ', 'DÍV. LÍQUIDA/PL', 'DÍV. LÍQUIDA/EBITDA', 'DÍV. LÍQUIDA/EBIT', 'PL/ATIVOS', 'PASSIVOS/ATIVOS', 'LIQ. CORRENTE', 'M. BRUTA', 'M. EBITDA', 'M. EBIT', 'M. LÍQUIDA', 'ROE', 'ROA', 'ROIC', 'GIRO ATIVOS', 'CAGR RECEITAS 5 ANOS', 'CAGR LUCROS 5 ANOS', 'Dividendo 2013', 'Dividendo 2014', 'Dividendo 2015', 'Dividendo 2016', 'Dividendo 2017', 'Dividendo 2018', 'Dividendo 2019', 'Dividendo 2020', 'Dividendo 2021', 'Dividendo 2022', 'Payout Médio', 'Payout Atual', 'Payout Menor Valor', 'Payout Maior Valor', 'Payout Médio 2018', 'Payout Médio 2019', 'Payout Médio 2020', 'Payout Médio 2021', 'Payout Médio 2022', 'Lucro Líquido 2018', 'Lucro Líquido 2019', 'Lucro Líquido 2020', 'Lucro Líquido 2021', 'Lucro Líquido 2022', 'Proventos 2018', 'Proventos 2019', 'Proventos 2020', 'Proventos 2021', 'Proventos 2022', 'Patriônio Líquido', 'Ativos', 'Ativo Circulante', 'Dívida Bruta', 'Disponibilidade', 'Dívida Líquida', 'Valor de Mercado', 'Valor de Firma', 'Nº Total de Papéis', 'Segmento de Listagem', 'Free Float']]
        # print(dado)

        # print(dado)

        dadosFormatados = json.dumps(dados, indent=2, ensure_ascii=False)

        # df.to_excel(nome_planilha, index=False)

        # dfFinal.to_excel(nome_planilha, index=False)



        #dados.values() = linhas da planilha
        # print(dadosFormatados)

        # listaTIKS = [i for i in dados.values()]

        # valores_adicionar = []

        # valores_adicionar.append(listaTIKS)


        return dados
    

# tickers = ['AZUL4', 'BBAS3', 'POMO4', 'CPLE6', 'CSAN3', 'PRIO3', 'ABEV3', 'KRSA3', 'PETZ3', 'LWSA3', 'SOMA3', 'PETR3', 'MRVE3', 'VBBR3', 'HBSA3', 'ENEV3', 'RAIL3', 'CMIN3', 'EQTL3', 'JBSS3', 'GOAU4', 'DMMO3', 'SUZB3', 'CCRO3', 'AERI3', 'GOLL4', 'RENT3', 'DXCO3', 'AMAR3', 'RRRP3', 'ASAI3', 'WEGE3', 'UGPA3', 'BEEF3', 'TOTS3', 'YDUQ3', 'TEND3', 'CEAB3', 'MULT3', 'GUAR3', 'SIMH3', 'ONCO3', 'TRPL4', 'AZEV4', 'ELET3', 'CMIG4', 'BBDC3', 'KLBN4', 'ALSO3', 'EMBR3', 'ECOR3', 'MOVI3', 'BBSE3', 'ESPA3', 'GMAT3', 'STBP3', 'VAMO3', 'TIMS3', 'MODL3', 'QUAL3', 'MEAL3', 'POSI3', 'RAPT4', 'CYRE3', 'ALPA4', 'GFSA3', 'LJQQ3', 'FLRY3', 'BRPR3', 'EZTC3', 'DIRR3', 'MYPK3', 'INEP3', 'LIGT3', 'BRSR6', 'PCAR3', 'ANIM3', 'RECV3', 'BRAP4', 'RADL3', 'HYPE3', 'LUPA3', 'BPAN4', 'CBAV3', 'MLAS3', 'GGPS3', 'ENBR3', 'MDIA3', 'SAPR4', 'CRFB3', 'SEQL3', 'SBSP3', 'RCSL4', 'BRKM5', 'WIZS3', 'OIBR4', 'JHSF3', 'AMBP3', 'SMTO3', 'GRND3', 'PSSA3', 'TASA4', 'BMGB4', 'RANI3', 'ODPV3', 'CXSE3', 'SLCE3', 'SHOW3', 'PGMN3', 'ENAT3', 'AESB3', 'VIVT3', 'OPCT3', 'EGIE3', 'CURY3', 'VIVA3', 'SEER3', 'TTEN3', 'EVEN3', 'ELET6', 'HBOR3', 'ARZZ3', 'SBFG3', 'TRIS3', 'DASA3', 'CAML3', 'CPFE3', 'JSLG3', 'NEOE3', 'CSMG3', 'BOAS3', 'MILS3', 'KEPL3', 'PORT3', 'AZEV3', 'AALR3', 'PTBL3', 'KLBN3', 'ENJU3', 'PARD3', 'TUPY3', 'BMOB3', 'VITT3', 'POMO3', 'INTB3', 'CPLE3', 'TPIS3', 'TRAD3', 'CMIG3']

tickers = []

try:
    tickers = pegar_tickers_planilha(nome_planilha)
except Exception:
    print("Erro")

lista_dict_acoes = []


for ticker in tickers:
    lista_dict_acoes.append(buscarTIK(ticker))
    print(f'Dados do Ticker {ticker} salvos!')

adicionar_ou_criar_planilha(lista_dict_acoes, nome_planilha)
organizar_planilha(nome_planilha)